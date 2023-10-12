# import openpyxl

# wb = openpyxl.load_workbook('output.xlsx')
# ws = wb.active

# for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
#     text = row[1]
#     print(text)
    # if i ==10:
    #     break



import time
import openpyxl
import openai
openai.api_key ="ENTER_YOUR_OPENAI_KEY"

wb = openpyxl.load_workbook('TextScrap.xlsx')
ws = wb.active

processed_urls = []
try:
    with open('processed_urls.txt', 'r') as f:
        processed_urls = f.read().splitlines()
except FileNotFoundError:
    pass

try:
    output_wb = openpyxl.load_workbook('chatgptResponse.xlsx')
    output_ws = output_wb.active
except FileNotFoundError:
    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active
    output_ws['A1'] = 'URL'
    output_ws['B1'] = 'Standard description en'
    output_ws['C1'] = 'Tittle'
    output_ws['D1'] = 'Standard bulletpoints en'
    output_ws['E1'] = 'Status'
    output_ws['F1'] = 'Duration (sec)'

for row in output_ws.iter_rows(min_row=2, values_only=True):
    url = row[0]
    processed_urls.append(url)

for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
    url = row[0]
    if url in processed_urls:
        print(f"Skipping URL {url} as it has already been processed.")
        continue
    
    text = row[1]
    
    if "Content not found" in text:
        print("Content not found")
        output_ws.cell(row=i+2, column=1, value=url)
        output_ws.cell(row=i+2, column=2, value="Content not found")
        output_ws.cell(row=i+2, column=3, value="Content not found")
        output_ws.cell(row=i+2, column=4, value="Content not found")
        output_ws.cell(row=i+2, column=5, value="Not Found")
        output_ws.cell(row=i+2, column=6, value=0)
        continue
    text = text.replace("\n", " ")
    
    print("URL:", url)

    start_time = time.time()
    prompt = "Rewrite this description with 500-1800 characters. It should be in HTML format where you should use strong text for subtitles such as Description, Features, Technical Specifications, and Packing List. Bullet points should also be used where suitable. Make sure to use suitable selling keywords used on Amazon. \n\n<b>Description:</b>\n{}\n\n<b>Features:</b>\n\n<b>Technical Specifications:</b>\n\n<b>Packing List:</b>\n".format(text)

    response = openai.Completion.create(
        engine="text-davinci-003",
        prompt=prompt,
        max_tokens=1800,
        n=1,
        stop=None,
        temperature=0.7,
        )
    
    product_description = response.choices[0].text.strip().replace("\n", "").replace("\t", "").replace("\r", "")
    print(product_description)
    
    title_prompt = "generate a selling title according to description which is suitable for the product. The title is not allowed to exceed 44 characters.: \n\n{}\n".format(product_description)
    
    title_response = openai.Completion.create(
        engine="text-davinci-003",
        prompt=title_prompt,
        max_tokens=50,
        n=1,
        stop=None,
        temperature=0.7,
    )

    generated_title = title_response.choices[0].text.strip().replace("\n", "").replace("\t", "").replace("\r", "")
    print(generated_title)
    
    
    keywords_prompt = "Generate only 5 unique selling keywords for this description. The summaries are not allowed to exceed 44 characters.Summaries should be  seprated by '|' not in html: \n\n{}\n".format(product_description)
    
    keywords_response = openai.Completion.create(
        engine="text-davinci-003",
        prompt=keywords_prompt,
        max_tokens=50,
        n=5,
        stop=None,
        temperature=0.7
    )
    generated_keywords = keywords_response.choices[0].text.strip().replace("\n", "").replace("\t", "").replace("\r", "")
    print(generated_keywords)
    
    
    
    output_ws.cell(row=i+2, column=1, value=url)
    output_ws.cell(row=i+2, column=2, value=product_description)
    output_ws.cell(row=i+2, column=3, value=generated_title)
    output_ws.cell(row=i+2, column=4, value=generated_keywords)
    output_ws.cell(row=i+2, column=5, value="Done")
    output_ws.cell(row=i+2, column=6, value=round(time.time() - start_time, 2))
    output_wb.save('chatgptResponse.xlsx')
    
    processed_urls.append(url)
    
    with open('processed_urls.txt', 'w') as f:
        for url in processed_urls:
            f.write(url + '\n')




