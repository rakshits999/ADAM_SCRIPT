import time
from selenium import webdriver
from bs4 import BeautifulSoup
import openpyxl

input_wb = openpyxl.load_workbook(filename='export_2023-04-08T01 08 22.717Z (1).xlsx')
input_ws = input_wb.active

output_wb = openpyxl.Workbook()
output_ws = output_wb.active

output_ws['A1'] = 'URL'
output_ws['B1'] = 'Scraped Text'
output_ws['C1'] = 'Status'


driver = webdriver.Chrome()

for i, row in enumerate(input_ws.iter_rows(min_row=2, values_only=True)):
    # if i == 3:
    #     break

    url = row[0]
    try:
        driver.get(url)
        scroll_pause_time = 1
        screen_height = driver.execute_script("return window.screen.height;")
        j = 1
        while True:
            scroll_height = driver.execute_script("return document.body.scrollHeight;")
            driver.execute_script(f"window.scrollTo(0, {screen_height * j});")
            time.sleep(scroll_pause_time)
            new_scroll_height = driver.execute_script("return document.body.scrollHeight;")
            if new_scroll_height == scroll_height:
                break

            j += 1
        time.sleep(1)
        html = driver.page_source
        soup = BeautifulSoup(html, "html.parser")
        content = soup.find("div", {"id": "product-description"})
        if content is None:
            content = soup.find("div", {"id": "module_product_specification"})

        if content:
            for element in content.select(".detailmodule_dynamic"):
                element.extract()
          
            style_tag = content.find("style")
            if style_tag:
                style_tag.extract()  
            text = ""
            for element in content.contents:
                if element.name != "style":
                    text += element.get_text()
    
            text = " ".join(text.split())
            print(text)
        


            output_ws.cell(row=i+2, column=1, value=url)
            output_ws.cell(row=i+2, column=2, value=text)
            output_ws.cell(row=i+2, column=3, value='Done')
            output_wb.save('TextScrap.xlsx')

        else:
            print("Content not found.")

            output_ws.cell(row=i+2, column=1, value=url)
            output_ws.cell(row=i+2, column=2, value='Content not found')
            output_ws.cell(row=i+2, column=3, value='Content not found')
            output_wb.save('TextScrap.xlsx')
    except Exception as e:
        print(f"Error occurred while processing URL: {url}. Error message: {e}")

driver.quit()



